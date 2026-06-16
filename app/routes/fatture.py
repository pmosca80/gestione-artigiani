import io
from datetime import datetime, date as _date

from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import get_current_user
from app import crud
from app.templates_config import templates
from app.services.audit import log_audit, get_actor, get_client_ip

router = APIRouter(prefix="/fatture", tags=["fatture"])


def _fmt_data(d) -> str:
    try:
        if hasattr(d, "strftime"):
            return d.strftime("%d/%m/%Y")
        return datetime.strptime(str(d)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return str(d) if d else ""


def _componi_email_sollecito(
    nome_cliente: str,
    numero_fmt: str,
    data_emissione: str,
    importo: float,
    nome_azienda: str,
    data_scadenza: str | None,
) -> str:
    scad_txt = ""
    if data_scadenza:
        try:
            scad_txt = f"(scadenza: <strong>{_fmt_data(data_scadenza)}</strong>) "
        except Exception:
            pass
    return f"""
    <html>
    <body style="font-family:'Segoe UI',Arial,sans-serif;background:#f8fafc;padding:32px 0;margin:0;">
    <div style="max-width:580px;margin:0 auto;background:white;border-radius:12px;
                box-shadow:0 2px 16px rgba(0,0,0,0.08);overflow:hidden;">
      <div style="background:linear-gradient(135deg,#0f172a,#1e3a5f);padding:28px 32px;">
        <p style="margin:0;color:#94a3b8;font-size:13px;">{nome_azienda}</p>
        <h1 style="margin:6px 0 0;color:white;font-size:20px;">Sollecito di pagamento</h1>
      </div>
      <div style="padding:28px 32px;">
        <p style="color:#374151;font-size:14px;line-height:1.7;margin:0 0 16px;">
          Gentile <strong>{nome_cliente}</strong>,
        </p>
        <p style="color:#374151;font-size:14px;line-height:1.7;margin:0 0 20px;">
          con la presente Le ricordiamo che la seguente fattura risulta ancora
          <strong>non saldata</strong>:
        </p>
        <div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:10px;
                    padding:16px 20px;margin-bottom:20px;">
          <table style="width:100%;font-size:14px;border-collapse:collapse;">
            <tr>
              <td style="color:#92400e;font-weight:600;padding:4px 0;">Fattura n°</td>
              <td style="text-align:right;font-weight:700;color:#111827;">{numero_fmt}</td>
            </tr>
            <tr>
              <td style="color:#92400e;font-weight:600;padding:4px 0;">Data emissione</td>
              <td style="text-align:right;color:#374151;">{_fmt_data(data_emissione)}</td>
            </tr>
            <tr>
              <td style="color:#92400e;font-weight:600;padding:8px 0 4px;">Importo dovuto</td>
              <td style="text-align:right;font-weight:700;font-size:18px;color:#dc2626;padding-top:8px;">
                € {importo:.2f}
              </td>
            </tr>
          </table>
        </div>
        <p style="color:#374151;font-size:14px;line-height:1.7;margin:0 0 16px;">
          La invitiamo cortesemente a provvedere al saldo {scad_txt}entro
          <strong>7 giorni</strong> dalla ricezione di questa comunicazione.
        </p>
        <p style="color:#374151;font-size:14px;line-height:1.7;margin:0;">
          Per qualsiasi chiarimento non esiti a contattarci.
          Cordiali saluti.
        </p>
      </div>
      <div style="padding:16px 32px;background:#f8fafc;border-top:1px solid #e5e7eb;">
        <p style="margin:0;font-size:12px;color:#6b7280;">
          <strong>{nome_azienda}</strong> — Comunicazione inviata tramite Mastro
        </p>
      </div>
    </div>
    </body>
    </html>
    """


@router.get("/", response_class=HTMLResponse)
def registro_fatture(
    request: Request,
    anno: int = None,
    inviata: int = None,
    sollecito_ok: int = None,
    errore: str = None,
    msg: str = None,
    link_creato: int = None,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.services.email import smtp_configurato
    from app.services.sdi import pec_configurata
    from app.services.stripe_service import stripe_configurato

    anni_disponibili = crud.get_anni_fatture(db, user_id)
    anno_sel = anno or (anni_disponibili[0] if anni_disponibili else datetime.now().year)

    fatture = crud.get_fatture_registro(db, user_id, anno_sel)

    tot_imponibile = sum(f.importo_imponibile or 0 for f in fatture)
    tot_iva = sum(f.importo_iva or 0 for f in fatture)
    tot_totale = sum(f.importo_totale or 0 for f in fatture)
    tot_da_incassare = sum(
        f.importo_totale or 0
        for f in fatture
        if (f.tipo_documento or "TD01") == "TD01"
        and f.lavoro
        and (f.lavoro.stato_pagamento or "da_pagare") not in ("pagato",)
    )

    azienda = crud.get_impostazioni_azienda(db, user_id)

    _ERRORI = {
        "email_mancante":       "Il cliente non ha un indirizzo email — aggiornalo nella scheda cliente.",
        "smtp_non_configurato": "Email non configurata. Imposta MAIL_USERNAME e MAIL_PASSWORD nelle variabili d'ambiente.",
        "pec_non_configurata":  "PEC non configurata. Vai a Impostazioni › Azienda e inserisci i dati PEC.",
        "fattura_non_trovata":  "Fattura non trovata.",
        "lavoro_non_trovato":   "Lavoro collegato non trovato.",
        "invio_fallito":        f"Invio fallito: {msg or 'errore sconosciuto'}",
        "sdi_fallito":          f"Invio a SDI fallito: {msg or 'errore sconosciuto'}",
        "pdf_fallito":          f"Invio PDF fallito: {msg or 'errore sconosciuto'}",
        "stripe_non_configurato": "Stripe non configurato. Aggiungi STRIPE_SECRET_KEY nelle variabili d'ambiente.",
        "stripe_fallito":       f"Creazione link Stripe fallita: {msg or 'errore sconosciuto'}",
    }

    return templates.TemplateResponse(
        request=request,
        name="fatture_registro.html",
        context={
            "fatture": fatture,
            "anni_disponibili": anni_disponibili,
            "anno_sel": anno_sel,
            "tot_imponibile": tot_imponibile,
            "tot_iva": tot_iva,
            "tot_totale": tot_totale,
            "tot_da_incassare": tot_da_incassare,
            "smtp_ok": smtp_configurato(),
            "pec_ok": pec_configurata(azienda) if azienda else False,
            "stripe_ok": stripe_configurato(),
            "flash_ok": inviata,
            "flash_sollecito": sollecito_ok,
            "flash_pdf": request.query_params.get("pdf_inviata"),
            "flash_link": request.query_params.get("link_creato"),
            "flash_err": _ERRORI.get(errore) if errore else None,
        },
    )


@router.get("/export-excel")
def export_excel(
    anno: int = None,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.services.piani import get_piano, ha_export
    if not ha_export(get_piano(db, user_id)):
        return RedirectResponse("/piani?upgrade=export", status_code=303)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    anni_disponibili = crud.get_anni_fatture(db, user_id)
    anno_sel = anno or (anni_disponibili[0] if anni_disponibili else datetime.now().year)
    fatture = crud.get_fatture_registro(db, user_id, anno_sel)
    azienda = crud.get_impostazioni_azienda(db, user_id)
    nome_az = azienda.nome_azienda or "Azienda"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fatture {anno_sel}"

    # Palette colori
    PURPLE = "7C3AED"
    PURPLE_LIGHT = "EDE9FE"
    GRAY_HEADER = "F8FAFC"
    BORDER_COLOR = "E5E7EB"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    # ── Riga titolo ───────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"Registro Fatture {anno_sel} — {nome_az}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=PURPLE)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"Generato il {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=10, color="9CA3AF")
    ws["A2"].alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16

    # ── Intestazioni colonne ──────────────────────────────────────────────────
    headers = ["N° Fattura", "Data", "Cliente", "P.IVA / CF Cliente",
               "Imponibile (€)", "IVA (€)", "Bollo (€)", "Totale (€)", "Stato", "File XML"]
    ws.append([])  # riga 3 vuota
    ws.append(headers)  # riga 4

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(name="Calibri", bold=True, size=10, color="374151")
        cell.fill = PatternFill("solid", fgColor=GRAY_HEADER.replace("#", ""))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 22

    # ── Righe dati ────────────────────────────────────────────────────────────
    tot_imponibile = 0.0
    tot_iva = 0.0
    tot_bollo = 0.0
    tot_totale = 0.0

    for f in fatture:
        lav = f.lavoro
        cli = lav.cliente if lav else None

        if cli:
            if cli.tipo_cliente == "azienda":
                nome_cli = cli.ragione_sociale or "—"
            else:
                nome_cli = f"{cli.nome or ''} {cli.cognome or ''}".strip() or "—"
            piva_cf = cli.partita_iva or cli.codice_fiscale or "—"
        else:
            nome_cli = "—"
            piva_cf = "—"

        numero_fmt = f"{f.anno}/{str(f.numero).zfill(3)}"
        imp = float(f.importo_imponibile or 0)
        iva = float(f.importo_iva or 0)
        bollo = float(f.importo_bollo or 0)
        tot = float(f.importo_totale or 0)
        tot_imponibile += imp
        tot_iva += iva
        tot_bollo += bollo
        tot_totale += tot

        row_data = [
            numero_fmt,
            f.data_emissione,
            nome_cli,
            piva_cf,
            imp,
            iva,
            bollo,
            tot,
            f.stato or "emessa",
            f.nome_file or "",
        ]
        ws.append(row_data)
        data_row = ws.max_row

        for col_idx in range(1, 11):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

        # Numero fattura in viola
        ws.cell(row=data_row, column=1).font = Font(name="Calibri", bold=True, size=10, color=PURPLE)
        ws.cell(row=data_row, column=1).fill = PatternFill("solid", fgColor=PURPLE_LIGHT)

        # Importi allineati a destra con formato valuta
        for col_idx in [5, 6, 7, 8]:
            cell = ws.cell(row=data_row, column=col_idx)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.row_dimensions[data_row].height = 18

    # ── Riga totali ───────────────────────────────────────────────────────────
    totals_row = ws.max_row + 1
    ws.cell(row=totals_row, column=1).value = f"TOTALI {anno_sel}"
    ws.cell(row=totals_row, column=1).font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ws.cell(row=totals_row, column=1).fill = PatternFill("solid", fgColor=PURPLE)
    ws.cell(row=totals_row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells(f"A{totals_row}:D{totals_row}")

    for col_idx, val in [(5, tot_imponibile), (6, tot_iva), (7, tot_bollo), (8, tot_totale)]:
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.value = val
        cell.number_format = '#,##0.00'
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin_border

    for col_idx in [2, 3, 4, 9, 10]:
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.fill = PatternFill("solid", fgColor=PURPLE)
        cell.border = thin_border

    ws.row_dimensions[totals_row].height = 22

    # ── Larghezze colonne ─────────────────────────────────────────────────────
    col_widths = [14, 12, 30, 20, 15, 12, 10, 15, 14, 28]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Output ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"fatture_{anno_sel}_{nome_az.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export.csv")
def export_csv_fatture(
    anno: int = None,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.services.piani import get_piano, ha_export
    if not ha_export(get_piano(db, user_id)):
        return RedirectResponse("/piani?upgrade=export", status_code=303)
    from fastapi.responses import Response as _Response
    anni_disponibili = crud.get_anni_fatture(db, user_id)
    anno_sel = anno or (anni_disponibili[0] if anni_disponibili else datetime.now().year)
    fatture_list = crud.get_fatture_registro(db, user_id, anno_sel)

    righe = ["N. Fattura,Data,Cliente,P.IVA/CF,Imponibile,IVA,Bollo,Totale,Stato"]
    for f in fatture_list:
        lav = f.lavoro
        cli = lav.cliente if lav else None
        if cli:
            nome_cli = (
                cli.ragione_sociale if cli.tipo_cliente == "azienda"
                else f"{cli.nome or ''} {cli.cognome or ''}".strip()
            ) or ""
            piva_cf = cli.partita_iva or cli.codice_fiscale or ""
        else:
            nome_cli = ""
            piva_cf = ""
        numero_fmt = f"{f.anno}/{str(f.numero).zfill(3)}"
        nome_cli = nome_cli.replace('"', "'")
        righe.append(
            f'{numero_fmt},{f.data_emissione},"{nome_cli}",{piva_cf},'
            f'{f.importo_imponibile or 0:.2f},{f.importo_iva or 0:.2f},'
            f'{f.importo_bollo or 0:.2f},{f.importo_totale or 0:.2f},{f.stato or "emessa"}'
        )

    content = "﻿" + "\n".join(righe)
    filename = f"fatture_{anno_sel}.csv"
    return _Response(
        content=content.encode("utf-8"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/liquidazione-iva", response_class=HTMLResponse)
def liquidazione_iva(
    request: Request,
    anno: int = None,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.models import VocePrimaNota, FatturaAcquisto
    _REGIMI_SENZA_IVA = {"RF19", "RF02"}
    # TD01=fattura, TD04=nota di credito (iva negativa), TD05=nota di debito (iva positiva)
    _TIPI_POSITIVI = {"TD01", "TD05"}
    _TIPI_NEGATIVI = {"TD04"}

    anni_disponibili = crud.get_anni_fatture(db, user_id)
    anno_sel = anno or (anni_disponibili[0] if anni_disponibili else datetime.now().year)

    fatture = crud.get_fatture_registro(db, user_id, anno_sel)
    azienda = crud.get_impostazioni_azienda(db, user_id)
    regime = (azienda.regime_fiscale if azienda else "RF01") or "RF01"
    forfettario = regime in _REGIMI_SENZA_IVA

    quartali: dict[int, dict] = {
        q: {"n": 0, "imponibile": 0.0, "iva_debito": 0.0, "iva_credito": 0.0, "mesi": label}
        for q, label in [
            (1, "Gen – Mar"), (2, "Apr – Giu"),
            (3, "Lug – Set"), (4, "Ott – Dic"),
        ]
    }

    # IVA a debito dalle fatture emesse
    for f in fatture:
        tipo_doc = f.tipo_documento or "TD01"
        if tipo_doc not in _TIPI_POSITIVI and tipo_doc not in _TIPI_NEGATIVI:
            continue
        try:
            mese = f.data_emissione.month if hasattr(f.data_emissione, "month") else int(str(f.data_emissione)[5:7])
        except Exception:
            continue
        q = (mese - 1) // 3 + 1
        segno = -1.0 if tipo_doc in _TIPI_NEGATIVI else 1.0
        quartali[q]["n"] += 1
        quartali[q]["imponibile"] += segno * float(f.importo_imponibile or 0)
        quartali[q]["iva_debito"] += segno * float(f.importo_iva or 0)

    # IVA a credito 1: voci prima nota (uscite con aliquota_iva > 0, tag manuale)
    voci_pn = (
        db.query(VocePrimaNota)
        .filter(
            VocePrimaNota.utente_id == user_id,
            VocePrimaNota.tipo == "uscita",
            VocePrimaNota.importo_iva > 0,
        )
        .all()
    )
    for v in voci_pn:
        try:
            data_str = str(v.data)
            anno_v = int(data_str[:4])
            mese_v = int(data_str[5:7])
        except Exception:
            continue
        if anno_v != anno_sel:
            continue
        q = (mese_v - 1) // 3 + 1
        quartali[q]["iva_credito"] += float(v.importo_iva or 0)

    # IVA a credito 2: fatture passive registrate (FatturaAcquisto)
    fatture_acq = (
        db.query(FatturaAcquisto)
        .filter(
            FatturaAcquisto.utente_id == user_id,
            FatturaAcquisto.anno == anno_sel,
            FatturaAcquisto.importo_iva > 0,
        )
        .all()
    )
    for fa in fatture_acq:
        try:
            mese_fa = fa.data_fattura.month if hasattr(fa.data_fattura, "month") else int(str(fa.data_fattura)[5:7])
        except Exception:
            continue
        q = (mese_fa - 1) // 3 + 1
        quartali[q]["iva_credito"] += float(fa.importo_iva or 0)

    # Calcola saldo per trimestre
    for q in quartali:
        quartali[q]["saldo"] = quartali[q]["iva_debito"] - quartali[q]["iva_credito"]

    tot_imponibile = sum(v["imponibile"] for v in quartali.values())
    tot_iva_debito = sum(v["iva_debito"] for v in quartali.values())
    tot_iva_credito = sum(v["iva_credito"] for v in quartali.values())
    tot_saldo = tot_iva_debito - tot_iva_credito

    return templates.TemplateResponse(
        request=request,
        name="liquidazione_iva.html",
        context={
            "anni_disponibili": anni_disponibili,
            "anno_sel": anno_sel,
            "quartali": quartali,
            "tot_imponibile": tot_imponibile,
            "tot_iva_debito": tot_iva_debito,
            "tot_iva_credito": tot_iva_credito,
            "tot_saldo": tot_saldo,
            "forfettario": forfettario,
            "regime": regime,
        },
    )


@router.post("/{fattura_id}/crea-link")
def crea_link_pagamento(
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.models import FatturaEmessa
    from app.services.stripe_service import crea_payment_link, stripe_configurato

    if not stripe_configurato():
        return RedirectResponse("/fatture/?errore=stripe_non_configurato", status_code=303)

    fattura = db.query(FatturaEmessa).filter(
        FatturaEmessa.id == fattura_id,
        FatturaEmessa.utente_id == user_id,
    ).first()
    if not fattura:
        return RedirectResponse("/fatture/?errore=fattura_non_trovata", status_code=303)

    try:
        numero_fmt = f"{fattura.anno}/{str(fattura.numero).zfill(3)}"
        link_id, link_url = crea_payment_link(
            numero_fmt=numero_fmt,
            importo_totale=float(fattura.importo_totale or 0),
            fattura_id=fattura.id,
            utente_id=user_id,
        )
        fattura.stripe_payment_link_id = link_id
        fattura.stripe_payment_link_url = link_url
        db.commit()
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&link_creato={fattura_id}",
            status_code=303,
        )
    except Exception as exc:
        import urllib.parse
        msg = urllib.parse.quote(str(exc)[:120])
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=stripe_fallito&msg={msg}",
            status_code=303,
        )


@router.post("/{fattura_id}/invia-email")
def invia_fattura_email(
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.models import FatturaEmessa
    from app.services.fatturapa import genera_xml_fatturapa, nome_file_fatturapa
    from app.services.email import invia_fattura_xml, smtp_configurato

    fattura = db.query(FatturaEmessa).filter(
        FatturaEmessa.id == fattura_id,
        FatturaEmessa.utente_id == user_id,
    ).first()
    if not fattura:
        return RedirectResponse("/fatture/?errore=fattura_non_trovata", status_code=303)

    lav = fattura.lavoro
    if not lav:
        return RedirectResponse("/fatture/?errore=lavoro_non_trovato", status_code=303)

    cli = lav.cliente
    if not cli or not (cli.email or "").strip():
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=email_mancante&id={fattura_id}",
            status_code=303,
        )

    if not smtp_configurato():
        return RedirectResponse("/fatture/?errore=smtp_non_configurato", status_code=303)

    azienda = crud.get_impostazioni_azienda(db, user_id)
    nome_file = fattura.nome_file or nome_file_fatturapa(azienda, lav)
    numero_fmt = f"{fattura.anno}/{str(fattura.numero).zfill(3)}"

    try:
        voci = crud.get_voci_preventivo(db, user_id, lav.id)
        xml_bytes = genera_xml_fatturapa(lav, cli, azienda, voci=voci or None)
        to_nome = (
            cli.ragione_sociale if cli.tipo_cliente == "azienda"
            else f"{cli.nome or ''} {cli.cognome or ''}".strip()
        ) or "Cliente"
        invia_fattura_xml(
            to_email=cli.email.strip(),
            to_nome=to_nome,
            from_nome=azienda.nome_azienda or "Azienda",
            numero_fattura=numero_fmt,
            data_emissione=fattura.data_emissione,
            importo_totale=float(fattura.importo_totale or 0),
            xml_bytes=xml_bytes,
            nome_file=nome_file,
        )
        fattura.stato = "inviata_sdi"
        db.commit()
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&inviata={fattura_id}",
            status_code=303,
        )
    except Exception as exc:
        import urllib.parse
        msg = urllib.parse.quote(str(exc)[:120])
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=invio_fallito&msg={msg}",
            status_code=303,
        )


@router.post("/{fattura_id}/invia-pdf")
def invia_fattura_pdf_route(
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Invia la fattura in PDF leggibile al cliente via email."""
    from app.models import FatturaEmessa
    from app.services.pdf_fattura import genera_pdf_fattura
    from app.services.email import invia_fattura_pdf, smtp_configurato

    if not smtp_configurato():
        return RedirectResponse("/fatture/?errore=smtp_non_configurato", status_code=303)

    fattura = db.query(FatturaEmessa).filter(
        FatturaEmessa.id == fattura_id,
        FatturaEmessa.utente_id == user_id,
    ).first()
    if not fattura:
        return RedirectResponse("/fatture/?errore=fattura_non_trovata", status_code=303)

    lav = fattura.lavoro
    if not lav:
        return RedirectResponse("/fatture/?errore=lavoro_non_trovato", status_code=303)

    cli = lav.cliente
    if not cli or not (cli.email or "").strip():
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=email_mancante",
            status_code=303,
        )

    azienda = crud.get_impostazioni_azienda(db, user_id)
    voci = crud.get_voci_preventivo(db, user_id, lav.id)
    numero_fmt = f"{fattura.anno}/{str(fattura.numero).zfill(3)}"
    to_nome = (
        cli.ragione_sociale if cli.tipo_cliente == "azienda"
        else f"{cli.nome or ''} {cli.cognome or ''}".strip()
    ) or "Cliente"

    try:
        pdf_bytes = genera_pdf_fattura(lav, cli, azienda, voci=voci or None)
        data_em = fattura.data_emissione
        data_str = (
            data_em.strftime("%d/%m/%Y") if hasattr(data_em, "strftime") else str(data_em)
        )
        anno = fattura.anno or (
            data_em.year if hasattr(data_em, "year") else _date.today().year
        )
        nome_file = f"fattura_{anno}_{str(fattura.numero).zfill(3)}.pdf"

        invia_fattura_pdf(
            to_email=cli.email.strip(),
            to_nome=to_nome,
            from_nome=azienda.nome_azienda or "Azienda",
            numero_fattura=numero_fmt,
            data_emissione=data_str,
            importo_totale=float(fattura.importo_totale or 0),
            pdf_bytes=pdf_bytes,
            nome_file=nome_file,
        )
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&pdf_inviata={fattura_id}",
            status_code=303,
        )
    except Exception as exc:
        import urllib.parse
        msg = urllib.parse.quote(str(exc)[:120])
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=pdf_fallito&msg={msg}",
            status_code=303,
        )


@router.post("/{fattura_id}/invia-sdi")
def invia_fattura_sdi(
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.models import FatturaEmessa
    from app.services.fatturapa import genera_xml_fatturapa, nome_file_fatturapa
    from app.services.sdi import invia_xml_a_sdi, pec_configurata
    import urllib.parse

    fattura = db.query(FatturaEmessa).filter(
        FatturaEmessa.id == fattura_id,
        FatturaEmessa.utente_id == user_id,
    ).first()
    if not fattura:
        return RedirectResponse("/fatture/?errore=fattura_non_trovata", status_code=303)

    lav = fattura.lavoro
    if not lav:
        return RedirectResponse("/fatture/?errore=lavoro_non_trovato", status_code=303)

    azienda = crud.get_impostazioni_azienda(db, user_id)
    if not azienda or not pec_configurata(azienda):
        return RedirectResponse("/fatture/?errore=pec_non_configurata", status_code=303)

    try:
        voci = crud.get_voci_preventivo(db, user_id, lav.id)
        xml_bytes = genera_xml_fatturapa(lav, lav.cliente, azienda, voci=voci or None)
        nome_file = fattura.nome_file or nome_file_fatturapa(azienda, lav)
        invia_xml_a_sdi(xml_bytes, nome_file, azienda)
        fattura.stato = "inviata_sdi"
        db.commit()
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&inviata={fattura_id}",
            status_code=303,
        )
    except Exception as exc:
        msg = urllib.parse.quote(str(exc)[:120])
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=sdi_fallito&msg={msg}",
            status_code=303,
        )


@router.post("/{fattura_id}/pagamento")
def aggiorna_pagamento(
    request: Request,
    fattura_id: int,
    stato: str = Form(...),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Segna la fattura come pagata / da pagare aggiornando il Lavoro collegato."""
    crud.aggiorna_pagamento_fattura(db, fattura_id, user_id, stato)
    attore_id, attore_username = get_actor(request, db)
    log_audit(db, user_id, attore_id, attore_username,
              "pagamento_fattura", "fatture_emesse", fattura_id,
              {"stato": stato}, get_client_ip(request))
    return RedirectResponse("/fatture/", status_code=303)


@router.post("/{fattura_id}/nota-credito")
def crea_nota_credito(
    request: Request,
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Crea una nota di credito (TD04) a storno totale della fattura originale."""
    nc = crud.crea_nota_credito(db, user_id, fattura_id)
    if not nc:
        return RedirectResponse("/fatture/?errore=fattura_non_trovata", status_code=303)
    attore_id, attore_username = get_actor(request, db)
    log_audit(db, user_id, attore_id, attore_username,
              "nota_credito", "fatture_emesse", fattura_id,
              {"nc_id": nc.id, "totale": float(nc.importo_totale or 0)},
              get_client_ip(request))
    return RedirectResponse(
        f"/fatture/?anno={nc.anno}&nc_creata={nc.id}", status_code=303
    )


@router.post("/{fattura_id}/sollecito-cliente")
def sollecito_cliente(
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Invia email di sollecito pagamento direttamente al cliente."""
    from app.models import FatturaEmessa
    from app.services.email import invia_email, smtp_configurato

    if not smtp_configurato():
        return RedirectResponse("/fatture/?errore=smtp_non_configurato", status_code=303)

    fattura = db.query(FatturaEmessa).filter(
        FatturaEmessa.id == fattura_id,
        FatturaEmessa.utente_id == user_id,
    ).first()
    if not fattura:
        return RedirectResponse("/fatture/?errore=fattura_non_trovata", status_code=303)

    lav = fattura.lavoro
    if not lav:
        return RedirectResponse("/fatture/?errore=lavoro_non_trovato", status_code=303)

    cli = lav.cliente
    if not cli or not (cli.email or "").strip():
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&errore=email_mancante",
            status_code=303,
        )

    azienda = crud.get_impostazioni_azienda(db, user_id)
    nome_az = (azienda.nome_azienda if azienda else "") or ""
    numero_fmt = f"{fattura.anno}/{str(fattura.numero).zfill(3)}"
    nome_cli = (
        cli.ragione_sociale if cli.tipo_cliente == "azienda"
        else f"{cli.nome or ''} {cli.cognome or ''}".strip()
    ) or "Cliente"

    oggetto = f"Sollecito pagamento — Fattura n. {numero_fmt}"
    corpo = _componi_email_sollecito(
        nome_cli, numero_fmt, fattura.data_emissione,
        float(fattura.importo_totale or 0), nome_az,
        lav.data_scadenza_pagamento,
    )

    if invia_email(cli.email.strip(), oggetto, corpo):
        return RedirectResponse(
            f"/fatture/?anno={fattura.anno}&sollecito_ok={fattura_id}",
            status_code=303,
        )
    return RedirectResponse(
        f"/fatture/?anno={fattura.anno}&errore=invio_fallito",
        status_code=303,
    )


@router.post("/{fattura_id}/stato")
def aggiorna_stato(
    request: Request,
    fattura_id: int,
    stato: str = Form(...),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    crud.aggiorna_stato_fattura(db, fattura_id, user_id, stato)
    attore_id, attore_username = get_actor(request, db)
    log_audit(db, user_id, attore_id, attore_username,
              "stato_fattura", "fatture_emesse", fattura_id,
              {"stato": stato}, get_client_ip(request))
    return RedirectResponse("/fatture/", status_code=303)


@router.post("/crea-da-lavoro/{lavoro_id}")
def crea_fattura_da_lavoro(
    request: Request,
    lavoro_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Crea FatturaEmessa nel registro senza forzare il download dell'XML."""
    from app.services.piani import get_piano, ha_fatturapa
    if not ha_fatturapa(get_piano(db, user_id)):
        return RedirectResponse("/piani?upgrade=fatturapa", status_code=303)

    from app.services.fatturapa import (
        errori_fatturapa, nome_file_fatturapa, bollo_dovuto, _REGIMI_SENZA_IVA,
    )

    lavoro = crud.get_lavoro_by_id(db, lavoro_id, user_id)
    if not lavoro:
        raise HTTPException(status_code=404)

    cliente = lavoro.cliente
    azienda = crud.get_impostazioni_azienda(db, user_id)

    errori = errori_fatturapa(lavoro, cliente, azienda)
    if errori:
        items = "".join(f"<li>{err}</li>" for err in errori)
        return HTMLResponse(
            status_code=422,
            content=f"""<!DOCTYPE html>
<html lang="it">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>FatturaPA — dati mancanti</title>
  <link rel="stylesheet" href="/static/style.css">
  <style>
    body {{ font-family: 'DM Sans', sans-serif; background: #f8fafc; }}
    .wrap {{ max-width: 640px; margin: 60px auto; padding: 0 20px; }}
    .card {{ background: white; border: 1px solid #fca5a5; border-radius: 16px; padding: 32px; }}
    h2 {{ font-size: 20px; font-weight: 700; color: #991b1b; margin: 0 0 8px; }}
    p  {{ color: #6b7280; font-size: 14px; margin: 0 0 20px; }}
    ul {{ color: #374151; font-size: 14px; padding-left: 20px; line-height: 1.9; margin: 0 0 28px; }}
    .btn {{ display: inline-block; padding: 10px 20px; background: #2563eb; color: white;
            border-radius: 9px; font-size: 14px; font-weight: 700; text-decoration: none; }}
    .btn-gray {{ background: #f3f4f6; color: #374151; margin-left: 8px; }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h2>Impossibile generare la FatturaPA</h2>
      <p>Correggi i seguenti dati prima di procedere:</p>
      <ul>{items}</ul>
      <a href="/lavori/{lavoro_id}/modifica" class="btn">✏️ Modifica lavoro</a>
      <a href="/lavori/{lavoro_id}" class="btn btn-gray">← Torna alla scheda</a>
    </div>
  </div>
</body>
</html>""",
        )

    if not lavoro.numero_fattura:
        anno_gen, numero_gen = crud.genera_numero_fattura(db, user_id)
        lavoro.numero_fattura = numero_gen
        if not lavoro.data_fattura:
            lavoro.data_fattura = _date.today()
        db.commit()
        db.refresh(lavoro)

    lavoro.stato_fattura = "emessa"
    db.commit()

    data_em = lavoro.data_fattura or _date.today()
    try:
        anno = data_em.year if hasattr(data_em, "year") else int(str(data_em)[:4])
    except (ValueError, TypeError):
        anno = _date.today().year

    imponibile_val = float(lavoro.importo_consuntivo or 0)
    regime_str = (azienda.regime_fiscale or "RF01").strip().upper()
    regime_senza_iva = regime_str in _REGIMI_SENZA_IVA
    aliquota_val = 0.0 if regime_senza_iva else float(lavoro.aliquota_iva if lavoro.aliquota_iva is not None else 22)
    iva_val = 0.0 if (regime_senza_iva or aliquota_val == 0) else float(
        lavoro.totale_iva or round(imponibile_val * aliquota_val / 100, 2)
    )
    bollo_val = bollo_dovuto(regime_senza_iva, imponibile_val)
    totale_val = imponibile_val if regime_senza_iva else float(lavoro.totale_documento or 0)
    totale_val = round(totale_val + bollo_val, 2)
    filename = nome_file_fatturapa(azienda, lavoro)

    fattura_salvata = crud.salva_fattura_emessa(
        db, user_id, lavoro.id, lavoro.numero_fattura, anno, data_em,
        imponibile_val, iva_val, totale_val, filename, azienda.regime_fiscale or "RF01",
        bollo=bollo_val,
    )
    attore_id, attore_username = get_actor(request, db)
    log_audit(db, user_id, attore_id, attore_username,
              "emette_fattura", "fatture_emesse", fattura_salvata.id,
              {"numero": lavoro.numero_fattura, "anno": anno,
               "imponibile": imponibile_val, "iva": iva_val, "totale": totale_val},
              get_client_ip(request))

    return RedirectResponse(f"/fatture/?anno={anno}&creata=1", status_code=303)


@router.get("/{fattura_id}/scarica-xml")
def scarica_xml_da_registro(
    fattura_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Genera e scarica l'XML di una FatturaEmessa già registrata, senza toccare il DB."""
    from app.services.piani import get_piano, ha_fatturapa
    if not ha_fatturapa(get_piano(db, user_id)):
        return RedirectResponse("/piani?upgrade=fatturapa", status_code=303)

    from app.models import FatturaEmessa as _FE
    from app.services.fatturapa import genera_xml_fatturapa, nome_file_fatturapa

    fattura = db.query(_FE).filter(_FE.id == fattura_id, _FE.utente_id == user_id).first()
    if not fattura:
        raise HTTPException(status_code=404)

    lav = fattura.lavoro
    if not lav:
        raise HTTPException(status_code=404)

    azienda = crud.get_impostazioni_azienda(db, user_id)
    voci = crud.get_voci_preventivo(db, user_id, lav.id)
    xml_bytes = genera_xml_fatturapa(
        lav, lav.cliente, azienda, voci=voci or None,
        tipo_documento=fattura.tipo_documento or "TD01",
        fattura_rif_numero=fattura.fattura_rif_numero,
        fattura_rif_anno=fattura.fattura_rif_anno,
    )
    filename = fattura.nome_file or nome_file_fatturapa(azienda, lav)

    return Response(
        content=xml_bytes,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
