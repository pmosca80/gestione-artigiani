from fastapi.responses import FileResponse
from pathlib import Path
from fastapi import UploadFile, File
from datetime import datetime

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from sqlalchemy.orm import Session
import json

from app.database import get_db
from app.dependencies import get_current_user, richiedi_titolare
from app import crud
from app.logger import get_logger
from app.limiter import user_limiter
from openpyxl import Workbook

logger = get_logger("admin")

from fastapi import HTTPException
from sqlalchemy import text as sql_text
from app.models import Utente, Cliente, Materiale, CaricoMateriale
from app.templates_config import templates
from app.validators import (
    NOME_MAX, RAGIONE_SOCIALE_MAX, PARTITA_IVA_MAX, CODICE_FISCALE_MAX,
    REGIME_FISCALE_MAX, INDIRIZZO_MAX, CAP_MAX, CITTA_MAX, PROVINCIA_MAX,
    TELEFONO_MAX, EMAIL_MAX, PEC_MAX, PASSWORD_MAX, clean, check_magic,
)


# ─── helpers import CSV/XLSX ────────────────────────────────────────────────

def _normalizza_chiave(k: str) -> str:
    import unicodedata
    k = k.strip().lower()
    k = unicodedata.normalize("NFD", k)
    return "".join(c for c in k if unicodedata.category(c) != "Mn")


_CLIENTI_MAP = {
    "tipo_cliente":    ["tipo", "tipo_cliente", "tipo cliente"],
    "nome":            ["nome", "first name", "firstname", "name"],
    "cognome":         ["cognome", "last name", "lastname", "surname"],
    "ragione_sociale": ["ragione sociale", "ragione_sociale", "azienda", "societa", "company", "denominazione"],
    "telefono":        ["telefono", "tel", "cellulare", "phone", "mobile"],
    "email":           ["email", "e-mail", "mail"],
    "indirizzo":       ["indirizzo", "via", "address"],
    "citta":           ["citta", "city", "comune", "localita"],
    "provincia":       ["provincia", "prov"],
    "cap":             ["cap", "zip", "codice postale"],
    "partita_iva":     ["partita iva", "partita_iva", "p.iva", "piva", "p. iva"],
    "codice_fiscale":  ["codice fiscale", "codice_fiscale", "cf"],
    "note":            ["note", "notes"],
}

_MATERIALI_MAP = {
    "nome":                    ["nome", "descrizione", "articolo", "prodotto", "item"],
    "categoria":               ["categoria", "category"],
    "unita_misura":            ["unita misura", "unita_misura", "um", "u.m.", "unita"],
    "quantita":                ["quantita", "qty", "qta", "pezzi", "quantita iniziale"],
    "scorta_minima":           ["scorta minima", "scorta_minima", "scorta", "minimo"],
    "prezzo_acquisto_pieno":   ["prezzo acquisto pieno", "prezzo acquisto", "costo acquisto", "prezzo_acquisto_pieno"],
    "prezzo_acquisto_scontato":["prezzo acquisto scontato", "prezzo_acquisto_scontato", "costo scontato"],
    "prezzo_vendita_default":  ["prezzo vendita", "prezzo_vendita", "prezzo_vendita_default", "prezzo"],
    "note":                    ["note", "notes"],
}


def _mappa_colonne(headers: list, mapping: dict) -> dict:
    norm = [_normalizza_chiave(h) for h in headers]
    out = {}
    for campo, aliases in mapping.items():
        for a in aliases:
            an = _normalizza_chiave(a)
            if an in norm:
                out[campo] = norm.index(an)
                break
    return out


def _leggi_file(contenuto: bytes, filename: str):
    import io, csv as _csv
    ext = Path(filename).suffix.lower()
    if ext == ".xlsx":
        import openpyxl as _opx
        wb = _opx.load_workbook(io.BytesIO(contenuto), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [str(c or "").strip() for c in rows[0]]
        data = [[str(v) if v is not None else "" for v in r] for r in rows[1:]]
        return headers, data
    else:
        text = None
        for enc in ("utf-8-sig", "latin-1"):
            try:
                text = contenuto.decode(enc)
                break
            except Exception:
                pass
        if text is None:
            text = contenuto.decode("latin-1", errors="replace")
        sample = text[:2048]
        try:
            dialect = _csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except Exception:
            dialect = _csv.excel
        reader = _csv.reader(io.StringIO(text), dialect)
        rows = list(reader)
        if not rows:
            return [], []
        headers = [c.strip() for c in rows[0]]
        return headers, rows[1:]

router = APIRouter(prefix="/impostazioni", tags=["impostazioni"], dependencies=[Depends(richiedi_titolare)])


@router.get("/azienda", response_class=HTMLResponse)
def form_impostazioni_azienda(request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user),):

    azienda = crud.get_impostazioni_azienda(db, user_id)

    return templates.TemplateResponse(
        request=request,
        name="impostazioni_azienda.html",
        context={"azienda": azienda}
    )


@router.post("/azienda")
def salva_impostazioni_azienda(
    request: Request,
    nome_azienda: str = Form("", max_length=RAGIONE_SOCIALE_MAX),
    partita_iva: str = Form("", max_length=PARTITA_IVA_MAX),
    codice_fiscale: str = Form("", max_length=CODICE_FISCALE_MAX),
    regime_fiscale: str = Form("RF01", max_length=REGIME_FISCALE_MAX),
    indirizzo: str = Form("", max_length=INDIRIZZO_MAX),
    cap: str = Form("", max_length=CAP_MAX),
    citta: str = Form("", max_length=CITTA_MAX),
    provincia: str = Form("", max_length=PROVINCIA_MAX),
    telefono: str = Form("", max_length=TELEFONO_MAX),
    email: str = Form("", max_length=EMAIL_MAX),
    pec_indirizzo: str = Form("", max_length=PEC_MAX),
    pec_smtp_host: str = Form("", max_length=200),
    pec_smtp_port: int = Form(465),
    pec_smtp_password: str = Form("", max_length=PASSWORD_MAX),
    aliquota_iva_default: str = Form("22"),
    invio_automatico_sdi: bool = Form(False),
    logo: UploadFile = File(None),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    logo_path = None

    if logo and logo.filename:
        estensione = Path(logo.filename).suffix.lower()
        if estensione in [".png", ".jpg", ".jpeg"]:
            from app.services.cloudinary_service import cloudinary_configurato, carica_immagine
            contenuto = logo.file.read()
            if check_magic(contenuto, estensione):
                if cloudinary_configurato():
                    logo_path = carica_immagine(contenuto, f"logo{estensione}", folder=f"loghi/{user_id}")
                else:
                    cartella = Path(f"uploads/loghi/{user_id}")
                    cartella.mkdir(parents=True, exist_ok=True)
                    percorso = cartella / f"logo{estensione}"
                    percorso.write_bytes(contenuto)
                    logo_path = str(percorso)

    crud.salva_impostazioni_azienda(
        db, user_id, nome_azienda, partita_iva, indirizzo, telefono, email, logo_path,
        codice_fiscale=codice_fiscale,
        regime_fiscale=regime_fiscale,
        cap=cap,
        citta=citta,
        provincia=provincia,
        pec_indirizzo=pec_indirizzo,
        pec_smtp_host=pec_smtp_host,
        pec_smtp_port=pec_smtp_port,
        pec_smtp_password=pec_smtp_password,
        aliquota_iva_default=float(aliquota_iva_default) if aliquota_iva_default else 22,
        invio_automatico_sdi=invio_automatico_sdi,
    )
    return RedirectResponse(url="/impostazioni/azienda?salvato=1", status_code=303)

@router.get("/export/clienti")
def esporta_clienti_excel(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):

    clienti = crud.get_clienti(db, utente_id=user_id)["items"]

    export_dir = Path("export")
    export_dir.mkdir(exist_ok=True)

    nome_file = f"clienti_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = export_dir / nome_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Clienti"

    ws.append([
        "ID",
        "Tipo",
        "Nome",
        "Cognome",
        "Ragione sociale",
        "Telefono",
        "Email",
        "Indirizzo",
        "Città",
        "Provincia",
        "CAP",
        "Residuo"
    ])

    for cliente in clienti:
        ws.append([
            cliente.id,
            cliente.tipo_cliente,
            cliente.nome,
            cliente.cognome,
            cliente.ragione_sociale,
            cliente.telefono,
            cliente.email,
            cliente.indirizzo,
            cliente.citta,
            cliente.provincia,
            cliente.cap,
            cliente.totale_residuo or 0
        ])

    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=nome_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@router.get("/export/lavori")
def esporta_lavori_excel(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):

    lavori = crud.get_lavori(db=db, utente_id=user_id)["items"]

    export_dir = Path("export")
    export_dir.mkdir(exist_ok=True)

    nome_file = f"lavori_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = export_dir / nome_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Lavori"

    ws.append([
        "ID",
        "Data lavoro",
        "Cliente",
        "Titolo",
        "Stato",
        "Priorità",
        "Preventivo",
        "Totale documento",
        "Pagato",
        "Residuo",
        "Margine",
        "Stato pagamento",
        "Scadenza pagamento"
    ])

    for lavoro in lavori:
        cliente = "-"
        if lavoro.cliente:
            cliente = f"{lavoro.cliente.nome or ''} {lavoro.cliente.cognome or ''}".strip()

        ws.append([
            lavoro.id,
            lavoro.data_lavoro,
            cliente,
            lavoro.titolo,
            lavoro.stato,
            lavoro.priorita,
            lavoro.importo_preventivato or 0,
            lavoro.totale_documento or 0,
            lavoro.importo_pagato or 0,
            lavoro.residuo_pagamento or 0,
            lavoro.margine or 0,
            lavoro.stato_pagamento,
            lavoro.data_scadenza_pagamento or ""
        ])

    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=nome_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@router.get("/export/materiali")
def esporta_materiali_excel(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):

    materiali = crud.get_materiali(db, user_id)

    export_dir = Path("export")
    export_dir.mkdir(exist_ok=True)

    nome_file = f"materiali_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = export_dir / nome_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Materiali"

    ws.append([
        "ID",
        "Nome",
        "Categoria",
        "Unità misura",
        "Quantità",
        "Scorta minima",
        "Prezzo acquisto pieno",
        "Prezzo acquisto scontato",
        "Prezzo vendita",
        "Valore magazzino",
        "Note"
    ])

    for materiale in materiali:
        valore_magazzino = (
            (materiale.quantita or 0)
            * (materiale.prezzo_acquisto_scontato or materiale.prezzo_acquisto_pieno or 0)
        )

        ws.append([
            materiale.id,
            materiale.nome,
            materiale.categoria,
            materiale.unita_misura,
            materiale.quantita or 0,
            materiale.scorta_minima or 0,
            materiale.prezzo_acquisto_pieno or 0,
            materiale.prezzo_acquisto_scontato or 0,
            materiale.prezzo_vendita_default or 0,
            valore_magazzino,
            materiale.note or ""
        ])

    wb.save(file_path)

    return FileResponse(
        path=file_path,
        filename=nome_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.get("/import", response_class=HTMLResponse)
def pagina_import(
    request: Request,
    tipo: str = None,
    inseriti: int = None,
    saltati: int = None,
    errori: int = None,
    user_id: int = Depends(get_current_user),
):
    return templates.TemplateResponse(
        request=request,
        name="import_dati.html",
        context={"tipo": tipo, "inseriti": inseriti, "saltati": saltati, "errori": errori},
    )


@router.post("/import/clienti")
async def importa_clienti(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    contenuto = await file.read()
    try:
        headers, righe = _leggi_file(contenuto, file.filename)
    except Exception:
        return RedirectResponse("/impostazioni/import?tipo=clienti&errori=1", status_code=303)

    colmap = _mappa_colonne(headers, _CLIENTI_MAP)

    email_esistenti = {
        r.email.lower()
        for r in db.query(Cliente.email).filter(
            Cliente.utente_id == user_id,
            Cliente.email.isnot(None),
            Cliente.email != "",
        ).all()
        if r.email
    }

    inseriti = saltati = errori = 0

    def _val(riga, campo):
        idx = colmap.get(campo)
        if idx is None or idx >= len(riga):
            return None
        v = str(riga[idx]).strip() if riga[idx] is not None else ""
        return v or None

    for riga in righe:
        try:
            nome = _val(riga, "nome") or ""
            cognome = _val(riga, "cognome") or ""
            ragione_sociale = _val(riga, "ragione_sociale") or ""
            if not nome and not cognome and not ragione_sociale:
                saltati += 1
                continue
            email = _val(riga, "email") or ""
            if email and email.lower() in email_esistenti:
                saltati += 1
                continue
            tipo_c = _val(riga, "tipo_cliente") or ("azienda" if ragione_sociale and not nome else "privato")
            c = Cliente(
                utente_id=user_id,
                tipo_cliente=tipo_c,
                nome=nome or None,
                cognome=cognome or None,
                ragione_sociale=ragione_sociale or None,
                telefono=_val(riga, "telefono"),
                email=email or None,
                indirizzo=_val(riga, "indirizzo"),
                citta=_val(riga, "citta"),
                provincia=_val(riga, "provincia"),
                cap=_val(riga, "cap"),
                partita_iva=_val(riga, "partita_iva"),
                codice_fiscale=_val(riga, "codice_fiscale"),
                note=_val(riga, "note"),
                data_creazione=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            db.add(c)
            if email:
                email_esistenti.add(email.lower())
            inseriti += 1
        except Exception:
            errori += 1

    if inseriti > 0:
        db.commit()

    return RedirectResponse(
        f"/impostazioni/import?tipo=clienti&inseriti={inseriti}&saltati={saltati}&errori={errori}",
        status_code=303,
    )


@router.post("/import/materiali")
async def importa_materiali(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    contenuto = await file.read()
    try:
        headers, righe = _leggi_file(contenuto, file.filename)
    except Exception:
        return RedirectResponse("/impostazioni/import?tipo=materiali&errori=1", status_code=303)

    colmap = _mappa_colonne(headers, _MATERIALI_MAP)

    nomi_esistenti = {
        r.nome.lower()
        for r in db.query(Materiale.nome).filter(
            Materiale.utente_id == user_id,
            Materiale.nome.isnot(None),
        ).all()
        if r.nome
    }

    inseriti = saltati = errori = 0

    def _val(riga, campo):
        idx = colmap.get(campo)
        if idx is None or idx >= len(riga):
            return None
        v = str(riga[idx]).strip() if riga[idx] is not None else ""
        return v or None

    def _float(riga, campo, default=0.0):
        v = _val(riga, campo)
        if not v:
            return default
        try:
            return float(str(v).replace(",", ".").replace(" ", ""))
        except Exception:
            return default

    for riga in righe:
        try:
            nome = _val(riga, "nome")
            if not nome:
                saltati += 1
                continue
            if nome.lower() in nomi_esistenti:
                saltati += 1
                continue
            quantita = _float(riga, "quantita", 0)
            prezzo_pieno = _float(riga, "prezzo_acquisto_pieno", 0)
            prezzo_scontato = _float(riga, "prezzo_acquisto_scontato", 0)
            prezzo_vendita = _float(riga, "prezzo_vendita_default", 0)
            m = Materiale(
                utente_id=user_id,
                nome=nome,
                categoria=_val(riga, "categoria") or "",
                unita_misura=_val(riga, "unita_misura") or "pz",
                quantita=quantita,
                scorta_minima=_float(riga, "scorta_minima", 0),
                prezzo_acquisto_pieno=prezzo_pieno,
                prezzo_acquisto_scontato=prezzo_scontato,
                prezzo_vendita_default=prezzo_vendita,
                note=_val(riga, "note") or "",
                data_creazione=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
            db.add(m)
            db.flush()
            if quantita > 0:
                carico = CaricoMateriale(
                    utente_id=user_id,
                    materiale_id=m.id,
                    quantita_iniziale=quantita,
                    quantita_residua=quantita,
                    prezzo_acquisto=prezzo_scontato or prezzo_pieno or 0,
                    prezzo_vendita_default=prezzo_vendita or 0,
                    note="Importazione",
                    data_carico=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                )
                db.add(carico)
            nomi_esistenti.add(nome.lower())
            inseriti += 1
        except Exception:
            errori += 1

    if inseriti > 0:
        db.commit()

    return RedirectResponse(
        f"/impostazioni/import?tipo=materiali&inseriti={inseriti}&saltati={saltati}&errori={errori}",
        status_code=303,
    )


@router.get("/sicurezza", response_class=HTMLResponse)
def pagina_sicurezza(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    utente = db.query(Utente).filter(Utente.id == user_id).first()
    totp_abilitato = bool(getattr(utente, "totp_abilitato", False))

    qr_b64 = None
    totp_uri = None
    nuovo_secret = None

    if not totp_abilitato:
        import pyotp, qrcode, io, base64
        secret = getattr(utente, "totp_secret", None) or pyotp.random_base32()
        if not getattr(utente, "totp_secret", None):
            utente.totp_secret = secret
            db.commit()
        totp_uri = pyotp.totp.TOTP(secret).provisioning_uri(
            name=utente.email or utente.username,
            issuer_name="Mastro",
        )
        qr = qrcode.QRCode(box_size=5, border=4)
        qr.add_data(totp_uri)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        qr_b64 = base64.b64encode(buf.getvalue()).decode()
        nuovo_secret = secret

    return templates.TemplateResponse(
        request=request,
        name="impostazioni_sicurezza.html",
        context={
            "totp_abilitato": totp_abilitato,
            "qr_b64": qr_b64,
            "nuovo_secret": nuovo_secret,
        },
    )


@router.post("/sicurezza/attiva-2fa")
def attiva_2fa(
    request: Request,
    codice: str = Form(...),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    import pyotp
    utente = db.query(Utente).filter(Utente.id == user_id).first()
    secret = getattr(utente, "totp_secret", None)
    if not secret:
        return RedirectResponse("/impostazioni/sicurezza?errore=no_secret", status_code=303)

    totp = pyotp.TOTP(secret)
    if not totp.verify(codice.strip(), valid_window=1):
        return RedirectResponse("/impostazioni/sicurezza?errore=codice_errato", status_code=303)

    utente.totp_abilitato = True
    db.commit()
    return RedirectResponse("/impostazioni/sicurezza?attivato=1", status_code=303)


@router.post("/sicurezza/disattiva-2fa")
def disattiva_2fa(
    request: Request,
    password: str = Form(...),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.security import verify_password
    utente = db.query(Utente).filter(Utente.id == user_id).first()
    if not verify_password(password, utente.password):
        return RedirectResponse("/impostazioni/sicurezza?errore=password_errata", status_code=303)

    utente.totp_abilitato = False
    utente.totp_secret = None
    db.commit()
    return RedirectResponse("/impostazioni/sicurezza?disattivato=1", status_code=303)


@router.get("/admin", response_class=HTMLResponse)
def pagina_admin(request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user)):
    utente_corrente = db.query(Utente).filter(Utente.id == user_id).first()

    if not utente_corrente or utente_corrente.username != "admin":
        raise HTTPException(status_code=403, detail="Accesso negato")

    utenti = db.query(Utente).all()

    oggi = datetime.now()

    for utente in utenti:
        if utente.data_registrazione:
            try:
                dr = utente.data_registrazione
                reg = dr if not isinstance(dr, str) else datetime.strptime(dr, "%Y-%m-%d").date()
                utente.giorni_rimasti = 30 - (oggi.date() - reg).days
            except:
                utente.giorni_rimasti = None
        else:
            utente.giorni_rimasti = None

    return templates.TemplateResponse(
        request=request,
        name="admin.html",
        context={"utenti": utenti}
    )


@router.post("/admin/attiva/{utente_id}")
def attiva_utente(utente_id: int, request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user)):
    utente_corrente = db.query(Utente).filter(Utente.id == user_id).first()

    if not utente_corrente or utente_corrente.username != "admin":
        raise HTTPException(status_code=403, detail="Accesso negato")

    utente = db.query(Utente).filter(Utente.id == utente_id).first()
    if utente:
        utente.attivo = 2
        db.commit()

    return RedirectResponse(url="/impostazioni/admin", status_code=303)


@router.post("/admin/disattiva/{utente_id}")
def disattiva_utente(utente_id: int, request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user)):
    utente_corrente = db.query(Utente).filter(Utente.id == user_id).first()

    if not utente_corrente or utente_corrente.username != "admin":
        raise HTTPException(status_code=403, detail="Accesso negato")

    utente = db.query(Utente).filter(Utente.id == utente_id).first()
    if utente:
        utente.attivo = 0
        db.commit()

    return RedirectResponse(url="/impostazioni/admin", status_code=303)


@router.post("/admin/elimina/{utente_id}")
def elimina_utente(utente_id: int, request: Request, db: Session = Depends(get_db), user_id: int = Depends(get_current_user)):
    utente_corrente = db.query(Utente).filter(Utente.id == user_id).first()

    if not utente_corrente or utente_corrente.username != "admin":
        raise HTTPException(status_code=403, detail="Accesso negato")

    if utente_id == user_id:
        raise HTTPException(status_code=400, detail="Non puoi eliminare te stesso")

    utente = db.query(Utente).filter(Utente.id == utente_id).first()
    if not utente:
        return RedirectResponse(url="/impostazioni/admin", status_code=303)

    uid = {"uid": utente_id}
    # Tabelle figlie che referenziano lavori (devono venire prima di lavori)
    db.execute(sql_text("DELETE FROM materiali_usati_lavoro WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM sessioni_lavoro WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM pagamenti_lavoro WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM foto_lavori WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM allegati_lavoro WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM documenti_pdf WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM voci_preventivo WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM fatture_emesse WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM garanzie WHERE utente_id = :uid"), uid)
    # Tabelle magazzino
    db.execute(sql_text("DELETE FROM movimenti_magazzino WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM carichi_materiale WHERE utente_id = :uid"), uid)
    # Tabelle che referenziano lavori.id (devono precedere la cancellazione dei lavori)
    db.execute(sql_text("DELETE FROM sal_lavoro WHERE lavoro_id IN (SELECT id FROM lavori WHERE utente_id = :uid)"), uid)
    db.execute(sql_text("DELETE FROM rapportini_lavoro WHERE lavoro_id IN (SELECT id FROM lavori WHERE utente_id = :uid)"), uid)
    db.execute(sql_text("DELETE FROM timesheet_collab WHERE lavoro_id IN (SELECT id FROM lavori WHERE utente_id = :uid)"), uid)
    # Entità principali
    db.execute(sql_text("DELETE FROM lavori WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM materiali WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM promemoria_clienti WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM clienti WHERE utente_id = :uid"), uid)
    # Tabelle di configurazione e dati extra
    db.execute(sql_text("DELETE FROM impostazioni_azienda WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM template_preventivi WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM prima_nota WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM listino_voci WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM push_subscriptions WHERE utente_id = :uid"), uid)
    db.execute(sql_text("DELETE FROM inviti_account WHERE titolare_id = :uid"), uid)
    # Sgancia collaboratori prima di eliminare il titolare
    db.execute(sql_text("UPDATE utenti SET titolare_id = NULL WHERE titolare_id = :uid"), uid)

    try:
        db.delete(utente)
        db.commit()
        logger.info(f"Admin: utente {utente_id} eliminato")
    except Exception as e:
        db.rollback()
        logger.error(f"Admin: errore eliminazione utente {utente_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Errore eliminazione: {e}")

    return RedirectResponse(url="/impostazioni/admin", status_code=303)


@router.post("/admin/piano/{utente_id}")
def cambia_piano_utente(
    utente_id: int,
    request: Request,
    piano: str = Form(...),
    pro_scadenza: str = Form(""),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    utente_corrente = db.query(Utente).filter(Utente.id == user_id).first()
    if not utente_corrente or utente_corrente.username != "admin":
        raise HTTPException(status_code=403, detail="Accesso negato")

    utente = db.query(Utente).filter(Utente.id == utente_id).first()
    if utente:
        utente.piano = piano
        utente.pro_scadenza = pro_scadenza.strip() or None
        db.commit()
        logger.info(f"Admin: piano utente {utente_id} → {piano} (scadenza: {pro_scadenza or 'nessuna'})")

    return RedirectResponse(url="/impostazioni/admin", status_code=303)


# ── PROFILO UTENTE ────────────────────────────────────────────────────────────

_ERRORI_PROFILO = {
    "admin_no_delete": "L'account admin non può essere cancellato.",
    "conferma_errata": "Hai inserito un testo sbagliato. Scrivi CANCELLA per confermare.",
}

@router.get("/profilo", response_class=HTMLResponse)
def form_profilo(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
    errore: str = None,
):
    utente = db.query(Utente).filter(Utente.id == user_id).first()
    msg_errore = _ERRORI_PROFILO.get(errore) if errore else None
    return templates.TemplateResponse(
        request=request, name="impostazioni_profilo.html",
        context={"utente": utente, "errore": msg_errore, "successo": None}
    )


@router.post("/profilo")
def salva_profilo(
    request: Request,
    email: str = Form(""),
    password_attuale: str = Form(""),
    nuova_password: str = Form(""),
    conferma_password: str = Form(""),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    from app.security import verify_password, hash_password
    utente = db.query(Utente).filter(Utente.id == user_id).first()
    errore = None
    successo = None

    email = email.strip().lower()
    if email:
        if "@" not in email or "." not in email.split("@")[-1]:
            errore = "Indirizzo email non valido."
        else:
            esistente = db.query(Utente).filter(Utente.email == email, Utente.id != user_id).first()
            if esistente:
                errore = "Email già in uso da un altro account."
            else:
                utente.email = email
                successo = "Email aggiornata."

    _pw_cambiata = False
    if not errore and nuova_password:
        if not password_attuale:
            errore = "Inserisci la password attuale per cambiarla."
        else:
            pw_ok = False
            try:
                pw_ok = verify_password(password_attuale, utente.password)
            except Exception:
                pass
            if not pw_ok and utente.password == password_attuale:
                pw_ok = True
            if not pw_ok:
                errore = "Password attuale errata."
            elif nuova_password != conferma_password:
                errore = "Le nuove password non coincidono."
            elif len(nuova_password) < 8:
                errore = "La nuova password deve essere di almeno 8 caratteri."
            else:
                utente.password = hash_password(nuova_password)
                _pw_cambiata = True
                successo = (successo + " Password aggiornata." if successo else "Password aggiornata.")

    if not errore:
        db.commit()
        if _pw_cambiata:
            request.session["pw_sig"] = utente.password[-12:]
        if not successo:
            successo = "Nessuna modifica effettuata."

    return templates.TemplateResponse(
        request=request, name="impostazioni_profilo.html",
        context={"utente": utente, "errore": errore, "successo": successo}
    )


@router.get("/gdpr/export")
@user_limiter.limit("5/minute")
def esporta_dati_gdpr(
    request: Request,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    """Export dati personali in formato machine-readable (art. 20 GDPR — portabilità)."""
    from app.services.gdpr import esporta_dati_utente
    dati = esporta_dati_utente(db, user_id)
    contenuto = json.dumps(dati, indent=2, ensure_ascii=False, default=str)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        content=contenuto.encode("utf-8"),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="dati_mastro_{ts}.json"'},
    )


@router.post("/cancella-account")
def cancella_account(
    request: Request,
    conferma_testo: str = Form(""),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    import secrets as _secrets
    utente = db.query(Utente).filter(Utente.id == user_id).first()
    if not utente or utente.username == "admin":
        return RedirectResponse("/impostazioni/profilo?errore=admin_no_delete", status_code=303)
    if conferma_testo.strip().upper() != "CANCELLA":
        return RedirectResponse("/impostazioni/profilo?errore=conferma_errata", status_code=303)

    from app.services.gdpr import cancella_dati_utente
    cancella_dati_utente(db, user_id)

    utente.email = None
    utente.attivo = 0
    utente.username = f"deleted_{user_id}_{_secrets.token_hex(4)}"
    utente.password = _secrets.token_hex(32)
    utente.token_verifica = None
    utente.token_reset = None
    utente.token_reset_scadenza = None
    utente.stripe_customer_id = None
    utente.stripe_subscription_id = None
    db.commit()

    request.session.clear()
    return RedirectResponse("/login?account_cancellato=1", status_code=303)